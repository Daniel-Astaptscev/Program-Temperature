from bs4 import BeautifulSoup
import openpyxl
import requests

# переменные м. р. и г. о. для url
regions = ['isakly', 'koshki-samarskaya',
           'elhovka-elhovskiy-rayon-samarskaya',
           'verhnie-belozerki',
           'bolshaya-chernigovka',
           'pestravka', 'bezenchuk',
           'hvorostyanka-samarskaya',
           'lopatino-volzhskiy-rayon-samarskaya',
           'malaya-malyshevka', 'neftegorsk',
           'kinel-cherkassy',
           'pohvistnevo', 'Pohvistnevo',
           'otradniy', 'privolzhe-samarskaya',
           'chapaevsk', 'novokuybishevsk',
           'samara', 'zhigulevsk',
           'oktyabrsk', 'bayderyakovo-samarskaya',
           'Neftegorsk', 'bogatoe-samarskaya']
# общий итоговый список по каждому региону
result_max, result_min, result_average = [], [], []
month_xlsx = {'январь': 2, 'февраль': 6, 'март': 10, 'апрель': 14,
              'сентябрь': 18, 'октябрь': 22, 'ноябрь': 26, 'декабрь': 30}


# проверка на ответ от сервера
def check_url(url):
    if url.status_code == 200:
        return 'Запрос к сайту выполнен => '
    else:
        return 'Программа обнаружила исключение: отсутствует доступ к странице сайта => '


# подсчёт среднего значения и добавление результата в общий список
def average_value(lst_max, lst_min):
    clear_temp = [[value.text for value in lst_max], [value.text for value in lst_min]]
    temp = [[int(value.replace('°', '')) for value in clear_temp[0]],
            [int(value.replace('°', '')) for value in clear_temp[1]]]

    try:
        total = round(sum(temp[0]) / len(temp[0]), 1)
        result_max.append(total)
        result_average.append(round((total - fact), 1))
        total = round(sum(temp[1]) / len(temp[1]), 1)
        result_min.append(total)
    except ZeroDivisionError:
        return 'программа обнаружила исключение: отсутствуют данные на сайте'
    else:
        return f'выгружено 100%'


month = input('За какой месяц выгрузить данные с сайта? ').lower()

while month not in month_xlsx:
    print('Введено неверное имя месяца, повторите попытку')
    month = input('За какой месяц выгрузить данные с сайта? ').lower()

year = input('За какой год внести данные в таблицу? ')
fact = float(input('Каким будет климатическое значение для расчёта за месяц? '))
print('Выполняется работа программы, пожалуйста подождите...')

for reg in regions:
    url = f'https://{reg}.nuipogoda.ru/{month}-{year}'
    page = requests.get(url)
    print(check_url(page), end='')

    # сбор числовых значений с сайта
    soup = BeautifulSoup(page.text, "html.parser")
    print(average_value(soup.find_all('span', class_='max'),
                        soup.find_all('span', class_='min')))

# сохранение данных в файл excel
temp_xlsx = openpyxl.load_workbook('./Temperature.xlsx')
sheet = temp_xlsx[year]

for num_row in range(3, 27):
    sheet.cell(row=num_row, column=month_xlsx[month]).value = result_max[
        num_row - 3]
    sheet.cell(row=num_row, column=month_xlsx[month] + 1).value = result_min[
        num_row - 3]
    sheet.cell(row=num_row, column=month_xlsx[month] + 2).value = fact
    sheet.cell(row=num_row, column=month_xlsx[month] + 3).value = \
        result_average[
            num_row - 3]
temp_xlsx.save('./Temperature.xlsx')
